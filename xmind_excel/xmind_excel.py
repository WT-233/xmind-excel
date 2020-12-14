# -*- coding: utf-8 -*-
"""
参考：[https://blog.csdn.net/JESSIE_liuym/article/details/104017884](https://blog.csdn.net/JESSIE_liuym/article/details/104017884)
#新增模块功能
#0820：修改了预期结果脏数据的bug
"""

"""
待优化：
1、预期结果取消判断空处理；
2、增加前置步骤（需要允许为空的状态）；
3、导入导出的文件存放在本地，非git目录下；
"""


from xmindparser import xmind_to_dict
import openpyxl
import logging
import logging.handlers
from logging.handlers import TimedRotatingFileHandler
from datetime import datetime
import datetime
import timeit

class xmind_to_csv():
    def __init__(self,excelName,UserName):
        self.excelName=excelName
        self.username=UserName
        self.filename = self.excelName + '.xlsx'
        self.workbook = openpyxl.Workbook()
        self.workbook.save(self.filename)
        self.workbook = openpyxl.load_workbook(self.filename) #打开表
        self.moudle=''

    def createSheet(self):
        self.worksheet = self.workbook.create_sheet(self.moudle, 0)
        tinme_num =timeit.timeit("sum(range(100))")
        print (tinme_num)
        return 0



    def my_log(self):
        '''创建对象的类方法'''
        #创建日志搜集器
        mylog = logging.getLogger('mylog')
        mylog.setLevel('DEBUG')
        #创建日志输出渠道
        sh=logging.StreamHandler()
        sh.setLevel('INFO')
        # 按时间进行轮转的收集器
        file_name = datetime.now().strftime("%Y-%m-%d") + '.log'
        fh= TimedRotatingFileHandler(file_name,encoding='utf8',when='h',interval=24,backupCount=3)
        fh.setLevel('DEBUG')
        #将日志输出渠道和日志收集器进行绑定
        mylog.addHandler(fh)
        mylog.addHandler(sh)
        #设置日志输出格式
        fot = '%(asctime)s-[%(filename)s-->line:%(lineno)d]-%(message)s'
        formatter=logging.Formatter(fot)
        #将日志输出格式与输出渠道进行绑定
        sh.setFormatter(formatter)
        fh.setFormatter(formatter)
        return mylog




    def numberLen(self, value):
        try:
            return len(value['topics'])
        except KeyError:
            return 0

    def xmind_title(self, value):
        """获取xmind标题内容"""
        return value['title']

    def writeExcel(self, row, case,moudle):
        #不存在模块对应的表
        if moudle!=self.moudle:
            self.moudle=moudle
            self.createSheet()
        sort = 1
        for key, value in case.items():
            self.worksheet.cell(row=1, column=sort).value = key
            self.worksheet.cell(row=row, column=sort).value = value
            sort = sort + 1
            self.workbook.save(self.filename)



    def readXmind(self, FileName):
        self.rowNum = 1  # 计算测试用例的条数
        # self.caseDict = {}
        self.XmindContent = xmind_to_dict(FileName)[0]['topic']
        self.XmindTitle = self.xmind_title(self.XmindContent)
        FeaturesNUM = self.numberLen(self.XmindContent)
        TestName=''  #模块+标题名称
        for i in range(FeaturesNUM):
            TestFeatures = self.numberLen(self.XmindContent['topics'][i])
            TestName = self.XmindContent['topics'][i]['title']

            #标题1.1 填写确认
            if TestFeatures == 0:
                print('【',TestName,'】模块缺少，导出失败')
                self.my_log().debug('标题缺少，导出失败')
            else:
                for j in range(TestFeatures):
                    TestPoints = self.numberLen(self.XmindContent['topics'][i]['topics'][j])
                    TestName += self.XmindContent['topics'][i]['topics'][j]['title']
                    # 标题1.2 填写确认
                    if TestPoints == 0:
                        print('【',TestName,'】标题缺少，导出失败')
                        self.my_log().debug('标题缺少，导出失败')
                    else:
                        TestName +=  self.XmindContent['topics'][i]['topics'][j]['title']
                        for k in range(TestPoints):
                            TestTypes=self.numberLen(self.XmindContent['topics'][i]['topics'][j]['topics'][k])
                            TestName += self.XmindContent['topics'][i]['topics'][j]['topics'][k]['title']

                            # 标题1.3 填写确认
                            if TestTypes == 0:
                                print('【',TestName, '】标题缺少，导出失败')
                                self.my_log().debug('标题缺少，导出失败')
                            else:
                                for m in range(TestTypes):
                                    xmind_TestStep = ""  # 存放用例步骤
                                    xmind_TestExcpect = ""  # 存放用例结果

                                    xmind_num = 0
                                    TestItems = self.numberLen(self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]) # 用例步骤
                                    TestName += self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['title']
                                    xmind_number=''
                                    #用例等级
                                    if 'makers' in list(self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m].keys()):
                                        if  self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['makers'][
                                            0] == 'priority-1':
                                            xmind_number = '高'
                                        elif self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['makers'][
                                            0] == 'priority-2':
                                            xmind_number = '中'
                                        elif self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['makers'][
                                            0] == 'priority-3':
                                            xmind_number = '低'

                                    # 用例步骤 填写确认
                                    if TestItems == 0:
                                        print('【', TestName,'】下用例步骤缺少，导出失败')
                                        self.my_log().debug('用例步骤缺少，导出失败')
                                    else:
                                        for case in range(TestItems):
                                            self.caseDict = {}

                                            self.caseDict['TestFeatures'] = self.XmindContent['topics'][i]['title']
                                            self.caseDict['CaseTitles'] = self.XmindContent['topics'][i]['topics'][j]['title'] + '/' + self.XmindContent['topics'][i]['topics'][j]['topics'][k]['title'] + '/' + self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['title']  # 用例标题：添加/按钮/保存
                                            self.caseDict['TestStep'] = self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['topics'][case]['title']  # 用例步骤
                                            try:
                                                self.caseDict['TestExcpect'] = self.XmindContent['topics'][i]['topics'][j]['topics'][k]['topics'][m]['topics'][case]['topics'][0]['title']    # 用例预期结果
                                            except Exception as e:
                                                pass

                                            #填充步骤+结果
                                            xmind_num += 1
                                            #填充用例步骤
                                            try:
                                                xmind_TestStep += "步骤" + str(xmind_num) + "：" + self.caseDict['TestStep'] + "；\n"
                                            except Exception as e:
                                                print(f'第{i+1}个功能模块下的第{j+1}个功能点中的第{k+1}个测试点中的第{m+1}测试项下 组装用例步骤出现异常')

                                                self.caseDict['TestExcpect'] = ""

                                            # 填充预期结果
                                            try:
                                                if self.caseDict['TestExcpect'] != "":
                                                    xmind_TestExcpect += "步骤" + str(xmind_num) + "：" + self.caseDict['TestExcpect'] + "；\n"
                                            except Exception as e:
                                                self.caseDict['TestExcpect'] = ""

                                    #填充标题+步骤+结果 进行写入
                                    xmind_dic = {}
                                    xmind_dic["用例目录"]=""
                                    xmind_dic["用例名称"] = self.caseDict['CaseTitles']
                                    xmind_dic["需求ID"]=""
                                    xmind_dic["前置条件"] =""
                                    xmind_dic["用例步骤"] = xmind_TestStep
                                    xmind_dic["预期结果"] = xmind_TestExcpect
                                    xmind_dic["用例类型"] =""
                                    xmind_dic["用例状态"] =""
                                    xmind_dic["用例等级"] = xmind_number
                                    xmind_dic["创建人"] =""
                                    xmind_dic["用例类型自定义"] =""
                                    xmind_moudle = self.caseDict['TestFeatures'] + '.xlsx'

                                    self.rowNum = self.rowNum + 1
                                    self.writeExcel(self.rowNum, xmind_dic,xmind_moudle)
            self.rowNum = 1

        tinme_num = timeit.timeit("sum(range(100))")
        print('readXmind:',tinme_num)

if __name__ == '__main__':
    UserName = "不使用"#input('UserName:')
    XmindFile = input('XmindFile:') #xmind_excel.xmind
    if XmindFile == 'wt':
        XmindFile = '广告导出计划.xmind'

    ExcelFile = input('ExcelFile:') #采销首页
    xmind_to_csv(ExcelFile,UserName).readXmind(XmindFile)
