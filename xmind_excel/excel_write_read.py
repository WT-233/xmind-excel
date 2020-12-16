# -*- coding: utf-8 -*-
"""
Excel 文件流处理
"""

import openpyxl
class ExcelWriteRead(object):
    def __init__(self, file_path, file_name):
        self.path = file_path + file_name +'.xlsx'
        self.file_name = file_name

        self.sheel_name = ''

        #初始化
        self.workbook = openpyxl.Workbook()
        self.worksheet = ""
        self.workbook.save(self.path)

    def excel_write(self, row, xmind_dic, sheel_name):
        # self.workbook = openpyxl.Workbook()

        if self.sheel_name != sheel_name:
            self.sheel_name = sheel_name
            self.worksheet = self.workbook.create_sheet(self.sheel_name, 0)

        sort = 1
        for key, value in xmind_dic.items():
            self.worksheet.cell(row=1, column=sort).value = key
            self.worksheet.cell(row=row, column=sort).value = value
            sort = sort + 1
            self.workbook.save(self.path)
