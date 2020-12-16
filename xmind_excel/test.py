import openpyxl
import xlwt
def operation_excel():

    xls = xlwt.Workbook()
    sht1 = xls.add_sheet("wt1")
    sht1.write(0,0,"jjjj")
    xls.save(r'/Users/wuting/wt测试.xls')

    # # fine_tune_no()
    # # 新建excel，改变sheet名字，并写入内容
    # wb = openpyxl.Workbook("/Users/wuting/")
    # ws = wb.active
    # ws.title = 'test_sheet1'
    # ws.cell(row=1, column=2).value = 1
    # wb.save('wt测试.xlsx')
    #
    # #加载excel，并创建sheet，并写入内容
    # path = r'no.xlsx'
    # wb = openpyxl.load_workbook(path)
    # ws1 = wb.create_sheet('name')
    # ws1.cell(row=1,column=2).value=3
    # wb.save(path)
    #
    #
    # # 加载excel，访问sheet，并写入内容
    # wb = openpyxl.load_workbook(path)
    # ws1 = wb['name']
    # ws1.cell(row=1,column=2).value=4
    # wb.save(path)
if __name__=='__main__':
    operation_excel()